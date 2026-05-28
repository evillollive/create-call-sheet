"""Tests for build_callsheet.py — xlsx workbook renderer."""
import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "create-call-sheet" / "scripts"))

import build_callsheet

SAMPLE_ANSWERS = Path(__file__).resolve().parent.parent / "create-call-sheet" / "examples" / "sample_answers.json"


@pytest.fixture
def sample_answers():
    return json.loads(SAMPLE_ANSWERS.read_text())


class TestBuild:
    def test_produces_xlsx(self, tmp_path, sample_answers):
        out = tmp_path / "test.xlsx"
        result = build_callsheet.build(sample_answers, out)
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_single_day_tabs(self, tmp_path, sample_answers):
        out = tmp_path / "test.xlsx"
        build_callsheet.build(sample_answers, out)
        wb = load_workbook(out)
        assert "Day 1" in wb.sheetnames
        assert "How to use" in wb.sheetnames

    def test_multi_day(self, tmp_path, sample_answers):
        day2 = dict(sample_answers["days"][0])
        day2["tab_name"] = "Day 2"
        day2["date_label"] = "Saturday, February 7, 2026"
        sample_answers["days"].append(day2)
        out = tmp_path / "test.xlsx"
        build_callsheet.build(sample_answers, out)
        wb = load_workbook(out)
        assert "Day 1" in wb.sheetnames
        assert "Day 2" in wb.sheetnames

    def test_no_days_raises(self, tmp_path):
        with pytest.raises(ValueError, match="days is required"):
            build_callsheet.build({"shared": {}, "days": []}, tmp_path / "x.xlsx")

    def test_minimal_answers(self, tmp_path):
        """Minimal valid answers — only required fields."""
        answers = {
            "shared": {"project_name": "Test"},
            "days": [{"crew_call": "9:00 AM"}],
        }
        out = tmp_path / "minimal.xlsx"
        result = build_callsheet.build(answers, out)
        assert result.exists()
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 2  # day tab + how to use

    def test_output_dir_created(self, tmp_path, sample_answers):
        out = tmp_path / "deep" / "nested" / "test.xlsx"
        result = build_callsheet.build(sample_answers, out)
        assert result.exists()

    def test_project_name_in_sheet(self, tmp_path, sample_answers):
        out = tmp_path / "test.xlsx"
        build_callsheet.build(sample_answers, out)
        wb = load_workbook(out)
        ws = wb["Day 1"]
        # Project name should be in row 4
        found = False
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell and "GitHub" in str(cell) and "ASOS" in str(cell):
                    found = True
                    break
        assert found, "Project name should appear in the sheet"
