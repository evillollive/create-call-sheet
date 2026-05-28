"""Render a call sheet workbook (.xlsx) from a JSON of answers.

The agent collects user input section by section, builds an `answers` dict
matching the schema in examples/sample_answers.json, then calls:

    python scripts/build_callsheet.py path/to/answers.json path/to/output.xlsx

One workbook is produced. For multi-day shoots, one tab per day plus a
"How to use" tab.

The visual design matches Call Sheet Template.xlsx (the user-approved spec).
Sections are emitted in fixed order. If a section's data is empty, it's skipped
to keep the sheet clean.
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Palette / style (matches template) ----------
INK        = "111111"
PAPER      = "FFFFFF"
SUBTLE     = "F4F4F5"
BAND       = "111111"
BAND_TEXT  = "FFFFFF"
ACCENT     = "FFD400"
ACCENT_INK = "111111"
RULE       = "D4D4D8"
DEPT_BG    = "27272A"
DEPT_FG    = "FFFFFF"
NOTE_BG    = "FEF3C7"
INPUT_BG   = "FAFAFA"
DEEMP_BG   = "6B6B6B"
DEEMP_FG   = "FFFFFF"
LINK_FG    = "2563EB"

FONT_NAME  = "Arial"
COLS       = 14


# ---------- Helper: typed cell writer ----------
def _font(size=None, bold=None, color=None, italic=None):
    kwargs = {"name": FONT_NAME}
    if size is not None: kwargs["size"] = size
    if bold is not None: kwargs["bold"] = bold
    if color is not None: kwargs["color"] = color
    if italic is not None: kwargs["italic"] = italic
    return Font(**kwargs)


def _align(h=None, v=None, wrap=None, indent=None):
    kwargs = {}
    if h is not None: kwargs["horizontal"] = h
    if v is not None: kwargs["vertical"] = v
    if wrap is not None: kwargs["wrap_text"] = wrap
    if indent is not None: kwargs["indent"] = indent
    return Alignment(**kwargs)


def _border():
    thin = Side(style="thin", color=RULE)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _put(ws, coord, value=None, *, fill=None, font=None, align=None, border=None, hyperlink=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", start_color=fill)
    if align: c.alignment = align
    if border: c.border = border
    if hyperlink: c.hyperlink = hyperlink
    return c


def _merge(ws, rng):
    ws.merge_cells(rng)


def _span(ws, row, c1, c2, value, *, fill=None, font=None, align=None, border=None, hyperlink=None):
    """Merge row across columns c1-c2 (inclusive) and write value into the left cell."""
    rng = f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}"
    _merge(ws, rng)
    return _put(ws, f"{get_column_letter(c1)}{row}", value,
                fill=fill, font=font, align=align, border=border, hyperlink=hyperlink)


def _band(ws, row, text, *, fill=BAND, fg=BAND_TEXT, size=11):
    _span(ws, row, 1, COLS, text,
          fill=fill,
          font=_font(size=size, bold=True, color=fg),
          align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[row].height = 22


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Day tab builder ----------
def _build_day_tab(wb, day_data: dict, shared: dict, tab_name: str, day_idx: int, day_total: int):
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 12, 16, 16, 14, 14, 14, 14, 16, 14, 14, 12, 12, 4])

    border = _border()

    # ----- Header band -----
    ws.row_dimensions[1].height = 8
    _span(ws, 2, 1, COLS, "CALL SHEET",
          fill=BAND, font=_font(size=22, bold=True, color=BAND_TEXT),
          align=_align(h="left", v="center", indent=2))
    ws.row_dimensions[2].height = 38
    ws.row_dimensions[3].height = 6

    # Project / Day / Date row
    _span(ws, 4, 1, 8, shared.get("project_name") or "[ Project Name ]",
          font=_font(size=16, bold=True, color=INK),
          align=_align(h="left", v="center"))
    _span(ws, 4, 9, 11, f"DAY {day_idx} OF {day_total}",
          fill=SUBTLE, font=_font(size=11, bold=True),
          align=_align(h="center", v="center"), border=border)
    _span(ws, 4, 12, 14, day_data.get("date_label") or "[ Day/Month/YYYY ]",
          fill=SUBTLE, font=_font(size=11, bold=True),
          align=_align(h="center", v="center"), border=border)
    ws.row_dimensions[4].height = 26

    # Client / Production Co / Agency / Job#
    labels = [("CLIENT", 1, 4, shared.get("client")),
              ("PRODUCTION CO.", 5, 8, shared.get("production_company")),
              ("AGENCY", 9, 11, shared.get("agency")),
              ("JOB #", 12, 14, shared.get("job_number"))]
    for lab, c1, c2, _ in labels:
        _span(ws, 5, c1, c2, lab,
              font=_font(size=9, bold=True, color="6B6B6B"),
              align=_align(h="left", v="center", indent=1), fill=PAPER)
    ws.row_dimensions[5].height = 14
    for _, c1, c2, val in labels:
        _span(ws, 6, c1, c2, val or "",
              fill=INPUT_BG, border=border,
              font=_font(size=11),
              align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[6].height = 22

    # ----- AT A GLANCE -----
    ws.row_dimensions[7].height = 10
    _band(ws, 8, "AT A GLANCE")

    # 3 big cards
    _merge(ws, "A9:D11"); _merge(ws, "E9:I11"); _merge(ws, "J9:N11")
    crew_call = day_data.get("crew_call") or "[ 00:00 AM/PM ]"
    loc_lines = day_data.get("location") or {}
    loc_text = "LOCATION\n" + "\n".join(filter(None, [
        loc_lines.get("address_line_1"),
        loc_lines.get("address_line_2"),
        loc_lines.get("city_state_zip"),
    ])) if loc_lines.get("address_line_1") else "LOCATION\n[ Address line 1 ]\n[ City, State ZIP ]"
    hosp = day_data.get("hospital") or {}
    hosp_text = "NEAREST HOSPITAL\n" + "\n".join(filter(None, [
        hosp.get("name"),
        hosp.get("address"),
        hosp.get("phone"),
    ])) + "\nDIAL 911 IN EMERGENCY" if hosp.get("name") else \
        "NEAREST HOSPITAL\n[ Hospital ]\n[ Address ]   •   [ Phone ]\nDIAL 911 IN EMERGENCY"

    _put(ws, "A9", f"CREW CALL\n{crew_call}",
         fill=ACCENT, font=_font(size=18, bold=True, color=ACCENT_INK),
         align=_align(h="center", v="center", wrap=True), border=border)
    _put(ws, "E9", loc_text,
         fill=PAPER, font=_font(size=12, bold=True, color=INK),
         align=_align(h="center", v="center", wrap=True), border=border)
    _put(ws, "J9", hosp_text,
         fill=PAPER, font=_font(size=11, bold=True, color=INK),
         align=_align(h="center", v="center", wrap=True), border=border)
    for rr in (9, 10, 11):
        ws.row_dimensions[rr].height = 26

    # Quick info ribbon
    ws.row_dimensions[12].height = 8
    ribbon = [
        ("FIRST SHOT", 1, 3, day_data.get("first_shot")),
        ("WRAP",       4, 5, day_data.get("wrap")),
        ("LUNCH",      6, 7, day_data.get("lunch")),
        ("SUNRISE",    8, 9, day_data.get("sunrise")),
        ("SUNSET",    10, 11, day_data.get("sunset")),
        ("WEATHER",   12, 14, day_data.get("weather")),
    ]
    for lab, c1, c2, _ in ribbon:
        _span(ws, 13, c1, c2, lab,
              font=_font(size=9, bold=True, color="6B6B6B"),
              align=_align(h="center", v="center"))
    ws.row_dimensions[13].height = 14
    for _, c1, c2, val in ribbon:
        _span(ws, 14, c1, c2, val or "",
              fill=INPUT_BG, border=border,
              font=_font(size=11, bold=True),
              align=_align(h="center", v="center"))
    ws.row_dimensions[14].height = 22

    # Weather lookup hyperlink
    city = (day_data.get("city_for_lookup") or "").replace(" ", "+")
    date_q = (day_data.get("date_label") or "").replace(" ", "+")
    weather_url = f"https://www.google.com/search?q=weather+{city}+{date_q}"
    _span(ws, 15, 12, 14, "Check weather →",
          font=_font(size=9, italic=True, color=LINK_FG),
          align=_align(h="center", v="center"),
          hyperlink=weather_url)
    ws.row_dimensions[15].height = 14

    # ----- LOGISTICS -----
    ws.row_dimensions[16].height = 8
    _band(ws, 17, "LOGISTICS")

    for c1, c2, lab in [(1, 7, "PARKING"), (8, 14, "BUILDING ACCESS / LOAD IN")]:
        _span(ws, 18, c1, c2, lab,
              font=_font(size=9, bold=True, color="6B6B6B"),
              align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[18].height = 14

    _merge(ws, "A19:G21"); _merge(ws, "H19:N21")
    parking_txt = day_data.get("parking") or "[ Garage name, address, validation instructions, street vs. lot, what to avoid ]"
    access_txt = day_data.get("building_access") or "[ Reception floor, security check-in, loading dock instructions, who to ask for ]"
    _put(ws, "A19", parking_txt,
         fill=INPUT_BG, border=border,
         font=_font(size=10),
         align=_align(h="left", v="top", wrap=True, indent=1))
    _put(ws, "H19", access_txt,
         fill=INPUT_BG, border=border,
         font=_font(size=10),
         align=_align(h="left", v="top", wrap=True, indent=1))
    for rr in (19, 20, 21):
        ws.row_dimensions[rr].height = 22

    # On-site emergency contacts
    _span(ws, 22, 1, COLS, "ON-SITE EMERGENCY CONTACTS",
          font=_font(size=9, bold=True, color="6B6B6B"),
          align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[22].height = 14

    # headers
    spans = [(1, 3, "NAME"), (4, 5, "ROLE"), (6, 7, "PHONE"),
             (8, 10, "NAME"), (11, 12, "ROLE"), (13, 14, "PHONE")]
    for c1, c2, lab in spans:
        _span(ws, 23, c1, c2, lab,
              fill=SUBTLE, border=border,
              font=_font(size=9, bold=True),
              align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[23].height = 16

    on_site = day_data.get("on_site_contacts") or [{}, {}]
    while len(on_site) < 2:
        on_site.append({})
    for i, rr in enumerate((24, 25)):
        contact = on_site[i] if i < len(on_site) else {}
        pairs = [
            (1, 3, contact.get("name", "")), (4, 5, contact.get("role", "")), (6, 7, contact.get("phone", "")),
        ]
        # Two contacts per row: first contact only fills left half, second contact fills right half
        # Actually our spec is one contact per row, two rows total
        for c1, c2, val in [(1, 3, contact.get("name", "")), (4, 5, contact.get("role", "")), (6, 7, contact.get("phone", ""))]:
            _span(ws, rr, c1, c2, val,
                  fill=INPUT_BG, border=border,
                  font=_font(size=11),
                  align=_align(h="left", v="center", indent=1))
        # Right side: leave blank for a second pair if user has more
        extra = on_site[i + 2] if (i + 2) < len(on_site) else {}
        for c1, c2, val in [(8, 10, extra.get("name", "")), (11, 12, extra.get("role", "")), (13, 14, extra.get("phone", ""))]:
            _span(ws, rr, c1, c2, val,
                  fill=INPUT_BG, border=border,
                  font=_font(size=11),
                  align=_align(h="left", v="center", indent=1))
        ws.row_dimensions[rr].height = 22

    # ----- SCHEDULE -----
    ws.row_dimensions[26].height = 8
    _band(ws, 27, "SCHEDULE")

    sched_headers = [(1, 2, "TIME"), (3, 7, "ACTIVITY"), (8, 10, "LOCATION"),
                     (11, 12, "TALENT"), (13, 14, "NOTES")]
    for c1, c2, lab in sched_headers:
        _span(ws, 28, c1, c2, lab,
              fill=DEPT_BG, border=border,
              font=_font(size=9, bold=True, color=DEPT_FG),
              align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[28].height = 18

    schedule = day_data.get("schedule") or []
    rows_needed = max(12, len(schedule))
    for i in range(rows_needed):
        rr = 29 + i
        item = schedule[i] if i < len(schedule) else {}
        cells = [
            (1, 2, item.get("time", "")),
            (3, 7, item.get("activity", "")),
            (8, 10, item.get("location", "")),
            (11, 12, item.get("talent", "")),
            (13, 14, item.get("notes", "")),
        ]
        striped = INPUT_BG if i % 2 == 0 else PAPER
        for c1, c2, val in cells:
            _span(ws, rr, c1, c2, val,
                  fill=striped, border=border,
                  font=_font(size=10),
                  align=_align(h="left", v="center", indent=1, wrap=True))
        ws.row_dimensions[rr].height = 20

    r = 29 + rows_needed

    # ----- CONTACTS -----
    ws.row_dimensions[r].height = 8
    r += 1
    _band(ws, r, "CONTACTS  •  CREW, CLIENT, TALENT, VENDORS")
    r += 1

    contact_headers = [(1, 3, "POSITION"), (4, 5, "NAME"), (6, 6, "PRONOUNS"),
                       (7, 8, "PHONE"), (9, 12, "EMAIL"), (13, 13, "CALL"), (14, 14, "WRAP")]
    for c1, c2, lab in contact_headers:
        if c1 == c2:
            _put(ws, f"{get_column_letter(c1)}{r}", lab,
                 fill=DEPT_BG, border=border,
                 font=_font(size=9, bold=True, color=DEPT_FG),
                 align=_align(h="left", v="center", indent=1))
        else:
            _span(ws, r, c1, c2, lab,
                  fill=DEPT_BG, border=border,
                  font=_font(size=9, bold=True, color=DEPT_FG),
                  align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[r].height = 18
    r += 1

    contacts_by_dept = shared.get("contacts_by_department") or {}
    department_order = [
        "PRODUCTION", "CAMERA", "G&E", "SOUND", "ART", "VANITIES",
        "SUPPORT", "CLIENT", "AGENCY", "TALENT / INTERVIEWEES", "VENDORS",
    ]
    for dept in department_order:
        rows = contacts_by_dept.get(dept) or []
        if not rows:
            continue
        # Dept header
        _span(ws, r, 1, COLS, dept,
              fill=DEPT_BG, font=_font(size=10, bold=True, color=DEPT_FG),
              align=_align(h="left", v="center", indent=1))
        ws.row_dimensions[r].height = 16
        r += 1
        for entry in rows:
            _span(ws, r, 1, 3, entry.get("position", ""),
                  fill=SUBTLE, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            _span(ws, r, 4, 5, entry.get("name", ""),
                  fill=INPUT_BG, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            _put(ws, f"F{r}", entry.get("pronouns", ""),
                 fill=INPUT_BG, border=border, font=_font(size=10),
                 align=_align(h="center", v="center"))
            _span(ws, r, 7, 8, entry.get("phone", ""),
                  fill=INPUT_BG, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            _span(ws, r, 9, 12, entry.get("email", ""),
                  fill=INPUT_BG, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            _put(ws, f"M{r}", entry.get("call", ""),
                 fill=INPUT_BG, border=border, font=_font(size=10),
                 align=_align(h="center", v="center"))
            _put(ws, f"N{r}", entry.get("wrap", ""),
                 fill=INPUT_BG, border=border, font=_font(size=10),
                 align=_align(h="center", v="center"))
            ws.row_dimensions[r].height = 19
            r += 1

    # ----- WARDROBE -----
    wardrobe = shared.get("wardrobe") or {}
    if wardrobe.get("recommended") or wardrobe.get("avoid"):
        ws.row_dimensions[r].height = 8; r += 1
        _band(ws, r, "WARDROBE & STYLING GUIDANCE (FOR TALENT)"); r += 1
        _span(ws, r, 1, 7, "RECOMMENDED",
              font=_font(size=9, bold=True, color="6B6B6B"),
              align=_align(h="left", v="center", indent=1))
        _span(ws, r, 8, 14, "AVOID",
              font=_font(size=9, bold=True, color="6B6B6B"),
              align=_align(h="left", v="center", indent=1))
        ws.row_dimensions[r].height = 14
        r += 1
        _merge(ws, f"A{r}:G{r+2}")
        _merge(ws, f"H{r}:N{r+2}")
        _put(ws, f"A{r}", wardrobe.get("recommended", ""),
             fill=INPUT_BG, border=border, font=_font(size=10),
             align=_align(h="left", v="top", wrap=True, indent=1))
        _put(ws, f"H{r}", wardrobe.get("avoid", ""),
             fill=INPUT_BG, border=border, font=_font(size=10),
             align=_align(h="left", v="top", wrap=True, indent=1))
        for rr in range(r, r + 3):
            ws.row_dimensions[rr].height = 22
        r += 3

    # ----- MEALS -----
    meals = shared.get("meals") or {}
    if any(meals.values()):
        ws.row_dimensions[r].height = 8; r += 1
        _band(ws, r, "MEALS  •  CRAFTY  •  ALLERGIES"); r += 1

        headers = [(1, 2, "MEAL"), (3, 6, "VENDOR"), (7, 8, "PHONE"),
                   (9, 9, "TIME"), (10, 14, "NOTES")]
        for c1, c2, lab in headers:
            if c1 == c2:
                _put(ws, f"{get_column_letter(c1)}{r}", lab,
                     fill=DEPT_BG, border=border,
                     font=_font(size=9, bold=True, color=DEPT_FG),
                     align=_align(h="left", v="center", indent=1))
            else:
                _span(ws, r, c1, c2, lab,
                      fill=DEPT_BG, border=border,
                      font=_font(size=9, bold=True, color=DEPT_FG),
                      align=_align(h="left", v="center", indent=1))
        ws.row_dimensions[r].height = 18
        r += 1

        for label in ("Breakfast", "Lunch", "Dinner", "Crafty / Snacks", "Water / Coffee"):
            entry = meals.get(label.lower().replace(" / ", "_").replace(" ", "_")) or {}
            _span(ws, r, 1, 2, label,
                  fill=SUBTLE, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            _span(ws, r, 3, 6, entry.get("vendor", ""),
                  fill=INPUT_BG, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            _span(ws, r, 7, 8, entry.get("phone", ""),
                  fill=INPUT_BG, border=border, font=_font(size=10),
                  align=_align(h="center", v="center"))
            _put(ws, f"I{r}", entry.get("time", ""),
                 fill=INPUT_BG, border=border, font=_font(size=10),
                 align=_align(h="center", v="center"))
            _span(ws, r, 10, 14, entry.get("notes", ""),
                  fill=INPUT_BG, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            ws.row_dimensions[r].height = 20
            r += 1

        allergies = meals.get("allergies") or "Allergies / dietary restrictions: please flag any to the Producer in advance."
        _span(ws, r, 1, COLS, allergies,
              fill=NOTE_BG, border=border,
              font=_font(size=10, italic=True),
              align=_align(h="left", v="center", indent=1))
        ws.row_dimensions[r].height = 20
        r += 1

    # ----- NOTES / RULES -----
    notes_block = shared.get("notes_block") or []
    if notes_block:
        ws.row_dimensions[r].height = 8; r += 1
        _band(ws, r, "NOTES  •  RULES  •  ETIQUETTE"); r += 1
        for note in notes_block:
            _span(ws, r, 1, COLS, f"•  {note}",
                  fill=PAPER, border=border,
                  font=_font(size=10),
                  align=_align(h="left", v="center", wrap=True, indent=1))
            ws.row_dimensions[r].height = 20
            r += 1

    # ----- INVOICING -----
    invoicing = shared.get("invoicing") or {}
    if any(invoicing.values()):
        ws.row_dimensions[r].height = 8; r += 1
        _band(ws, r, "INVOICING"); r += 1
        rows_to_emit = [
            ("Job Reference (include on all invoices)", invoicing.get("job_reference")),
            ("Send invoices to (email)", invoicing.get("send_to")),
            ("Billing address", invoicing.get("billing_address")),
            ("Reimbursable / per diem notes", invoicing.get("notes")),
        ]
        for lab, val in rows_to_emit:
            _span(ws, r, 1, 5, lab,
                  fill=SUBTLE, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            _span(ws, r, 6, COLS, val or "",
                  fill=INPUT_BG, border=border, font=_font(size=10),
                  align=_align(h="left", v="center", indent=1))
            ws.row_dimensions[r].height = 22
            r += 1

    # ----- PRODUCTION REPORT (always emitted, de-prioritized) -----
    ws.row_dimensions[r].height = 12; r += 1
    _span(ws, r, 1, COLS,
          "PRODUCTION REPORT  (fill in during/after shoot — optional)",
          fill=DEEMP_BG, font=_font(size=10, bold=True, color=DEEMP_FG),
          align=_align(h="left", v="center", indent=1))
    ws.row_dimensions[r].height = 18
    r += 1
    pr_fields = [
        ("1st AM Shot", "1st Meal Start", "1st Meal End"),
        ("1st PM Shot", "2nd Meal Start", "2nd Meal End"),
        ("Camera Wrap", "Crew Wrap", "Location Closed"),
    ]
    for a, b, c in pr_fields:
        _span(ws, r, 1, 2, a,
              fill=SUBTLE, border=border,
              font=_font(size=9, color="3F3F46"),
              align=_align(h="left", v="center", indent=1))
        _span(ws, r, 3, 5, "",
              fill=INPUT_BG, border=border, font=_font(size=10),
              align=_align(h="center", v="center"))
        _span(ws, r, 6, 7, b,
              fill=SUBTLE, border=border,
              font=_font(size=9, color="3F3F46"),
              align=_align(h="left", v="center", indent=1))
        _span(ws, r, 8, 9, "",
              fill=INPUT_BG, border=border, font=_font(size=10),
              align=_align(h="center", v="center"))
        _span(ws, r, 10, 11, c,
              fill=SUBTLE, border=border,
              font=_font(size=9, color="3F3F46"),
              align=_align(h="left", v="center", indent=1))
        _span(ws, r, 12, COLS, "",
              fill=INPUT_BG, border=border, font=_font(size=10),
              align=_align(h="center", v="center"))
        ws.row_dimensions[r].height = 20
        r += 1
    ws.row_dimensions[r].height = 12

    # Print + freeze
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.print_options.horizontalCentered = True
    ws.freeze_panes = "A7"


def _build_howto_tab(wb):
    ws = wb.create_sheet("How to use")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110

    def w(row, text, *, size=11, bold=False, color=INK, height=18):
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = height

    w(2, "HOW TO USE THIS WORKBOOK", size=18, bold=True, height=28)
    w(3, "This call sheet was generated by the Create Call Sheet skill.", height=22)
    w(4, "Each shoot day has its own tab (Day 1, Day 2, etc.).")
    w(5, "")
    w(6, "EDITING", size=13, bold=True, height=24)
    w(7, "All gray cells are fill-ins. Yellow cells are highlighted call-outs.")
    w(8, "Anything in [ brackets ] is a placeholder — replace with shoot info.")
    w(9, "")
    w(10, "PRINTING", size=13, bold=True, height=24)
    w(11, "Pages are set to fit-to-width, portrait. Adjust in Print Preview if needed.")
    w(12, "")
    w(13, "WEATHER / SUNRISE", size=13, bold=True, height=24)
    w(14, "Sunrise and sunset are pre-filled where possible. Weather is left blank — click the 'Check weather' link in the quick-info ribbon to look it up.")


# ---------- Entry point ----------
def build(answers: dict, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    shared = answers.get("shared") or {}
    days = answers.get("days") or []
    if not days:
        raise ValueError("answers.days is required (at least one day)")

    wb = Workbook()
    # Remove default sheet; we'll create our own tabs in order.
    default = wb.active
    wb.remove(default)

    for i, day in enumerate(days, start=1):
        tab_name = day.get("tab_name") or (f"Day {i}" if len(days) > 1 else "Call Sheet")
        _build_day_tab(wb, day, shared, tab_name, i, len(days))

    _build_howto_tab(wb)

    wb.save(output_path)
    return output_path


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        sys.stderr.write("Usage: build_callsheet.py <answers.json> <output.xlsx>\n")
        return 2
    answers = json.loads(Path(argv[0]).read_text())
    out = build(answers, argv[1])
    print(json.dumps({"output": str(out)}))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
