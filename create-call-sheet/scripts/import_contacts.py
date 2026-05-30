"""Extract crew/client/talent contact entries from a past call sheet.

Accepts .xlsx (the format we produce) or .pdf (from other producers' sheets).
Returns a JSON blob the agent can review with the user before merging into the
local profile via profile.py.

Heuristic, not magic. The PDF parser scans for common patterns (Position/Name/
Phone/Email columns or labeled rows) and surfaces what it finds. The agent
should always show the results to the user and ask which to keep.

Usage:
    python import_contacts.py path/to/past_call_sheet.pdf
    python import_contacts.py path/to/past_call_sheet.xlsx

Output is a JSON object:
    {
      "source": "...",
      "candidates": [
        {"role": "DP", "name": "...", "phone": "...", "email": "..."}
      ]
    }
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path


PHONE_RE = re.compile(
    r"(\+?\d[\d\-.\s()]{7,}\d)"
)
EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")

# Role keywords we watch for. Each must match as a whole word/phrase to avoid
# false positives like "pa" matching "parking" or "pages".
ROLE_HINTS = [
    "director", "producer", "ep", "executive producer", "line producer",
    "production manager", "field producer", "dp", "director of photography",
    "1st ac", "2nd ac", "ac", "focus puller", "gaffer", "key grip", "grip",
    "swing", "best boy", "electrician", "sound mixer", "sound", "boom op", "boom",
    "hmu", "hair & makeup", "hair and makeup", "hair", "makeup", "wardrobe stylist", "stylist",
    "art director", "production designer", "set dresser",
    "production assistant", "pa", "driver", "teleprompter",
    "fixer", "local producer", "casting",
    "interviewee", "host", "presenter",
    "creative director", "account manager",
]

# Lines containing these substrings are section headers or boilerplate, not contacts.
_HEADER_NOISE = [
    "call sheet", "at a glance", "logistics", "parking", "building access",
    "schedule", "wardrobe & styling", "meals", "crafty", "allergies",
    "notes", "rules", "etiquette", "invoicing", "production report",
    "how to use", "talent / interviewees", "contacts", "crew",
    "fit-to-width", "check weather", "print preview",
    "recommended", "avoid", "confidentiality", "social media",
    "closed set", "health & safety", "respect the location",
    "sustainability", "on-site emergency",
]

# Build word-boundary patterns for role matching (longer phrases first)
_ROLE_PATTERNS = sorted(ROLE_HINTS, key=len, reverse=True)
_ROLE_RE = re.compile(
    r"\b(" + "|".join(re.escape(r) for r in _ROLE_PATTERNS) + r")\b",
    re.IGNORECASE,
)

# Roles that should always be uppercase abbreviations
_UPPERCASE_ROLES: dict[str, str] = {
    "dp": "DP", "director of photography": "DP",
    "pa": "PA", "production assistant": "PA",
    "hmu": "HMU", "hair & makeup": "HMU", "hair and makeup": "HMU",
    "ac": "AC", "1st ac": "1st AC", "2nd ac": "2nd AC",
    "ep": "EP", "executive producer": "EP",
    "g&e": "G&E",
}


def normalize_role(raw: str) -> str:
    """Normalize a role string: uppercase abbreviations, title-case everything else."""
    key = raw.strip().lower()
    if key in _UPPERCASE_ROLES:
        return _UPPERCASE_ROLES[key]
    return raw.strip().title()


def _is_header_noise(line: str) -> bool:
    """Return True if the line looks like a section header, not a contact."""
    low = line.lower()
    # Must have a real person indicator: phone, email, or a pipe-separated name
    has_phone = bool(PHONE_RE.search(line))
    has_email = bool(EMAIL_RE.search(line))
    # If no phone or email, check if the line matches known header patterns
    if not has_phone and not has_email:
        for noise in _HEADER_NOISE:
            if noise in low:
                return True
    # Very short lines with no contact info are likely headers
    if len(line) < 20 and not has_phone and not has_email:
        return True
    return False


def _scan_text(text: str, source: str) -> list[dict]:
    """Walk lines and try to pair role + name + phone + email."""
    out = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue

        # Skip header/boilerplate lines
        if _is_header_noise(line):
            continue

        m = _ROLE_RE.search(line)
        if not m:
            continue

        hit_role = m.group(1).strip()
        phones = PHONE_RE.findall(line)
        emails = EMAIL_RE.findall(line)

        # Must have at least a phone or email to be a real contact
        if not phones and not emails:
            continue

        # Extract name: for pipe-delimited lines (our xlsx output), split on pipes
        name = None
        parts = [p.strip() for p in line.split("|")]
        if len(parts) >= 3:
            # Find which part is the role, take the next part as the name
            for i, part in enumerate(parts):
                if _ROLE_RE.search(part):
                    if i + 1 < len(parts):
                        candidate = parts[i + 1].strip()
                        # Skip if it looks like a phone or email
                        if not PHONE_RE.match(candidate) and not EMAIL_RE.match(candidate):
                            name = candidate
                    break
        else:
            # Fallback: extract text between role and first phone/email
            try:
                after_role = line[m.end():]
                after_role = after_role.lstrip(" :|-—\t")
                cutoff = len(after_role)
                for p in phones:
                    idx = after_role.find(p)
                    if idx >= 0:
                        cutoff = min(cutoff, idx)
                for e in emails:
                    idx = after_role.find(e)
                    if idx >= 0:
                        cutoff = min(cutoff, idx)
                name = after_role[:cutoff].strip(" |\t-—")
            except Exception:
                pass

        # Clean up name: strip parenthetical pronouns
        if name:
            name = re.sub(r"\([^)]*\)", "", name).strip(" |\t-—")

        entry = {
            "role": normalize_role(hit_role),
            "name": name or "",
            "phone": phones[0] if phones else "",
            "email": emails[0] if emails else "",
            "source_line": line[:200],
        }
        if entry["name"] or entry["phone"] or entry["email"]:
            out.append(entry)
    return out


def _from_pdf(path: Path) -> list[dict]:
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        sys.stderr.write("pdfplumber required. Install with: pip install pdfplumber\n")
        sys.exit(2)
    text_parts = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            text_parts.append(t)
    return _scan_text("\n".join(text_parts), str(path))


def _from_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        sys.stderr.write("openpyxl required.\n")
        sys.exit(2)
    wb = load_workbook(str(path), data_only=True)
    lines = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
    return _scan_text("\n".join(lines), str(path))


def extract(path: str | Path) -> dict:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    if p.suffix.lower() == ".pdf":
        cands = _from_pdf(p)
    elif p.suffix.lower() in (".xlsx", ".xlsm"):
        cands = _from_xlsx(p)
    else:
        raise ValueError(f"Unsupported file type: {p.suffix}")

    # De-dupe by (role, name) pair
    seen = set()
    deduped = []
    for c in cands:
        key = (c.get("role", "").lower(), c.get("name", "").lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(c)

    return {"source": str(p), "candidates": deduped}


def main(argv: list[str]) -> int:
    if not argv:
        sys.stderr.write("Usage: import_contacts.py <file.pdf|.xlsx>\n")
        return 2
    print(json.dumps(extract(argv[0]), indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
